import customtkinter as ctk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill
import pdfplumber
import re
import os
import threading
import time

# Ρυθμίσεις Εμφάνισης
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class ModernDataMatcherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Αντιστοίχιση Δεδομένων Excel & PDF")
        self.root.geometry("1100x800")

        self.excel_path = ctk.StringVar()
        self.pdf_path = ctk.StringVar()
        self.export_matches_var = ctk.BooleanVar(value=True)
        self.highlight_excel_var = ctk.BooleanVar(value=True)
        self.search_code_var = ctk.StringVar()
        
        # Επιλογές Αναζήτησης
        self.regex_mode_var = ctk.StringVar(value="Έξυπνη Μετατροπή")
        self.custom_regex_var = ctk.StringVar(value="")
        self.final_regex = ""

        self.cached_excel_codes = []
        self.cached_pdf_codes = []
        self.last_loaded_excel = ""
        self.last_loaded_pdf = ""

        self.setup_ui()
        # Αρχικοποίηση του Regex
        self.on_regex_change(self.regex_mode_var.get())

    def setup_ui(self):
        main_container = ctk.CTkFrame(self.root, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=20)

        left_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

        # ΜΠΑΡΑ ΤΙΤΛΟΥ & ΚΟΥΜΠΙ ΟΔΗΓΙΩΝ 
        top_left_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        top_left_frame.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(top_left_frame, text="Εργαλεία Ελέγχου", font=("Arial", 20, "bold")).pack(side="left")
        ctk.CTkButton(top_left_frame, text="ℹ️ Οδηγίες Χρήσης", command=self.show_help, width=120, fg_color="#E3A82B", hover_color="#C08B1F", text_color="black").pack(side="right")

        # Επιλογή Αρχείων
        file_frame = ctk.CTkFrame(left_frame)
        file_frame.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(file_frame, text="1. Επιλογή Αρχείων", font=("Arial", 14, "bold")).pack(anchor="w", padx=15, pady=(10, 5))

        excel_row = ctk.CTkFrame(file_frame, fg_color="transparent")
        excel_row.pack(fill="x", padx=15, pady=5)
        ctk.CTkLabel(excel_row, text="Αρχείο Excel:").pack(side="left", padx=(0, 10))
        ctk.CTkEntry(excel_row, textvariable=self.excel_path, width=280, state='readonly').pack(side="left", padx=(0, 10))
        ctk.CTkButton(excel_row, text="Αναζήτηση", command=self.select_excel, width=90).pack(side="left")

        pdf_row = ctk.CTkFrame(file_frame, fg_color="transparent")
        pdf_row.pack(fill="x", padx=15, pady=(5, 15))
        ctk.CTkLabel(pdf_row, text="Αρχείο PDF:  ").pack(side="left", padx=(0, 10))
        ctk.CTkEntry(pdf_row, textvariable=self.pdf_path, width=280, state='readonly').pack(side="left", padx=(0, 10))
        ctk.CTkButton(pdf_row, text="Αναζήτηση", command=self.select_pdf, width=90).pack(side="left")

        # Κανόνας Αναζήτησης 
        regex_frame = ctk.CTkFrame(left_frame)
        regex_frame.pack(fill="x", pady=(0, 15))
        ctk.CTkLabel(regex_frame, text="2. Μορφή Κωδικού (Αναζήτηση)", font=("Arial", 14, "bold")).pack(anchor="w", padx=15, pady=(10, 5))

        self.regex_dropdown = ctk.CTkOptionMenu(
            regex_frame,
            variable=self.regex_mode_var,
            values=[
                "Έξυπνη Μετατροπή",
                "DYPA... με 3 παύλες",
                "Απλό Αλφαριθμητικό 5-10 χαρ",
                "Custom Regex (Για Προγραμματιστές)"
            ],
            command=self.on_regex_change
        )
        self.regex_dropdown.pack(fill="x", padx=15, pady=5)

        self.custom_regex_entry = ctk.CTkEntry(regex_frame, textvariable=self.custom_regex_var, text_color="lightgreen", placeholder_text="π.χ. ΑΔΥΓΥΕΦΓ-33242432")
        self.custom_regex_entry.pack(fill="x", padx=15, pady=5)
        self.custom_regex_entry.bind("<KeyRelease>", self.update_dynamic_regex)

        self.active_regex_display = ctk.CTkLabel(regex_frame, text="", font=("Consolas", 12), text_color="gray")
        self.active_regex_display.pack(anchor="w", padx=15, pady=(0, 10))

        # Μαζικός Έλεγχος & Επιλογές 
        options_frame = ctk.CTkFrame(left_frame)
        options_frame.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(options_frame, text="3. Μαζικός Έλεγχος", font=("Arial", 14, "bold")).pack(anchor="w", padx=15, pady=(10, 5))

        ctk.CTkCheckBox(options_frame, text="Δημιουργία Excel ΜΟΝΟ με τα κοινά", variable=self.export_matches_var).pack(anchor="w", padx=15, pady=5)
        ctk.CTkCheckBox(options_frame, text="Πράσινο Highlight στα κοινά (Στο αρχικό Excel)", variable=self.highlight_excel_var).pack(anchor="w", padx=15, pady=(5, 15))

        self.run_btn = ctk.CTkButton(options_frame, text="Εκτέλεση Ελέγχου", command=self.start_matching_thread, font=("Arial", 14, "bold"), fg_color="#2FA572", hover_color="#106A43")
        self.run_btn.pack(pady=(0, 10))

        self.loading_frame = ctk.CTkFrame(options_frame, height=60, fg_color="transparent")
        self.loading_frame.pack(fill="x", pady=5)
        self.loading_frame.pack_propagate(False) 

        self.progress = ctk.CTkProgressBar(self.loading_frame, mode="indeterminate", width=380)
        self.progress.set(0)
        self.status_label = ctk.CTkLabel(self.loading_frame, text="", text_color="gray")
        
        self.open_file_btn = ctk.CTkButton(options_frame, text="Άνοιγμα Αρχείου Κοινών", command=self.open_matches_file, fg_color="#1f538d")

        # --- ΠΛΑΙΣΙΟ 4: Μεμονωμένη Αναζήτηση ---
        search_frame = ctk.CTkFrame(left_frame)
        search_frame.pack(fill="x", pady=(20, 0))
        ctk.CTkLabel(search_frame, text="4. Μεμονωμένη Αναζήτηση Εργαζόμενου", font=("Arial", 14, "bold")).pack(anchor="w", padx=15, pady=(10, 5))

        s_row = ctk.CTkFrame(search_frame, fg_color="transparent")
        s_row.pack(fill="x", padx=15, pady=(5, 10))
        ctk.CTkLabel(s_row, text="Εισάγετε Κωδικό:").pack(side="left", padx=(0, 10))
        ctk.CTkEntry(s_row, textvariable=self.search_code_var, width=180).pack(side="left", padx=(0, 10))
        ctk.CTkButton(s_row, text="Αναζήτηση", command=self.start_search_thread, width=90).pack(side="left")
        
        self.search_result_label = ctk.CTkLabel(search_frame, text="", font=("Arial", 14, "bold"))
        self.search_result_label.pack(pady=(0, 15))

        # ====== ΔΕΞΙΑ ΣΤΗΛΗ (Προεπισκόπηση) ======
        right_frame = ctk.CTkFrame(main_container)
        right_frame.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        ctk.CTkLabel(right_frame, text="Προεπισκόπηση Κοινών (Matches)", font=("Arial", 16, "bold")).pack(pady=(15, 10))
        
        self.preview_box = ctk.CTkTextbox(right_frame, wrap="none", font=("Consolas", 14))
        self.preview_box.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        self.preview_box.insert("0.0", "Κάντε εκτέλεση ελέγχου για να\nεμφανιστούν τα αποτελέσματα εδώ...")
        self.preview_box.configure(state="disabled")

    # ΟΔΗΓΙΕΣ ΧΡΗΣΗΣ 
    def show_help(self):
        help_win = ctk.CTkToplevel(self.root)
        help_win.title("ℹ️ Οδηγίες Χρήσης & Βοήθεια")
        help_win.geometry("650x650")
        help_win.attributes("-topmost", True)
        
        scroll_frame = ctk.CTkScrollableFrame(help_win, fg_color="transparent")
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(scroll_frame, text="Καλώς ήρθατε στο File Comparison", font=("Arial", 22, "bold"), text_color="#2FA572").pack(anchor="w", pady=(0, 20))

        ctk.CTkLabel(scroll_frame, text=" ΒΗΜΑ 1: Επιλογή Αρχείων", font=("Arial", 16, "bold")).pack(anchor="w", pady=(10, 5))
        ctk.CTkLabel(scroll_frame, text="Επιλέξτε το αρχείο Excel (υποθέτουμε ότι οι κωδικοί βρίσκονται στην Πρώτη Στήλη) και το PDF που θέλετε να ελέγξετε.", font=("Arial", 14), justify="left", wraplength=580).pack(anchor="w", padx=10)

        ctk.CTkLabel(scroll_frame, text=" ΒΗΜΑ 2: Μορφή Κωδικού (Σημαντικό!)", font=("Arial", 16, "bold")).pack(anchor="w", pady=(20, 5))
        step2_text = (
            "Επιλέξτε από το μενού πώς είναι γραμμένοι οι κωδικοί σας. Αν οι 2 έτοιμοι τύποι δεν σας καλύπτουν, "
            "χρησιμοποιήστε την «Έξυπνη Μετατροπή».\n\n"
            "Απλά γράψτε έναν τυχαίο κωδικό της μορφής που ψάχνετε (π.χ. ΑΔΥΓΥΕΦΓ-33242432) και το "
            "πρόγραμμα θα «καταλάβει» αυτόματα τον κανόνα αναζήτησης για όλο το PDF!"
        )
        ctk.CTkLabel(scroll_frame, text=step2_text, font=("Arial", 14), justify="left", wraplength=580).pack(anchor="w", padx=10)

        ctk.CTkLabel(scroll_frame, text=" ΒΗΜΑ 3: Έλεγχος & Αποτελέσματα", font=("Arial", 16, "bold")).pack(anchor="w", pady=(20, 5))
        step3_text = (
            "Επιλέξτε αν θέλετε να παραχθεί ένα νέο αρχείο αποκλειστικά με τα κοινά (Matches) "
            "ή να χρωματιστούν με πράσινο χρώμα στο αρχικό σας Excel."
        )
        ctk.CTkLabel(scroll_frame, text=step3_text, font=("Arial", 14), justify="left", wraplength=580).pack(anchor="w", padx=10)

        warning_frame = ctk.CTkFrame(scroll_frame, fg_color="#5c1919", corner_radius=8)
        warning_frame.pack(fill="x", pady=(20, 20), padx=10)
        warning_text = "⚠️ ΠΡΟΣΟΧΗ: Αν επιλέξετε χρωματισμό (Highlight), βεβαιωθείτε ότι το αρχείο Excel είναι ΚΛΕΙΣΤΟ κατά την εκτέλεση. Σε αντίθετη περίπτωση, τα Windows θα εμποδίσουν την αποθήκευση."
        ctk.CTkLabel(warning_frame, text=warning_text, font=("Arial", 14, "bold"), text_color="#FFB3B3", justify="left", wraplength=550).pack(padx=15, pady=15)

        ctk.CTkLabel(scroll_frame, text=" Μεμονωμένη Αναζήτηση", font=("Arial", 16, "bold")).pack(anchor="w", pady=(5, 5))
        ctk.CTkLabel(scroll_frame, text="Χρησιμοποιήστε το κάτω πλαίσιο του παραθύρου για να ψάξετε ταχύτατα έναν συγκεκριμένο εργαζόμενο για άμεση επαλήθευση.", font=("Arial", 14), justify="left", wraplength=580).pack(anchor="w", padx=10)

    # ΜΗΧΑΝΙΣΜΟΣ REGEX & ΠΑΡΑΔΕΙΓΜΑΤΩΝ 
    def on_regex_change(self, choice):
        # Ενεργοποιούμε το πεδίο προσωρινά για να μπορέσουμε να γράψουμε το παράδειγμα
        self.custom_regex_entry.configure(state="normal")
        
        if choice == "DYPA... με 3 παύλες":
            self.final_regex = r"\b[A-Z]+-\d+-\d+-\d+\b"
            # Βάζουμε ένα fix παράδειγμα
            self.custom_regex_var.set("π.χ. DYPAUE-10031649-20250714-165727")
            # Κλειδώνουμε το πεδίο (το κείμενο γίνεται γκριζωπό)
            self.custom_regex_entry.configure(state="disabled")
            self.active_regex_display.configure(text=f"Ενεργός Κανόνας: {self.final_regex}")
            
        elif choice == "Απλό Αλφαριθμητικό 5-10 χαρ":
            self.final_regex = r"\b[A-Z0-9]{5,10}\b"
            # Βάζουμε ένα fix παράδειγμα
            self.custom_regex_var.set("π.χ. AB12345")
            # Κλειδώνουμε το πεδίο
            self.custom_regex_entry.configure(state="disabled")
            self.active_regex_display.configure(text=f"Ενεργός Κανόνας: {self.final_regex}")
            
        elif choice == "Custom Regex (Για Προγραμματιστές)":
            self.custom_regex_entry.configure(placeholder_text="Γράψτε Regex (π.χ. \\b\\d{4}\\b)")
            self.custom_regex_var.set(r"\b[A-Z]+-\d+\b")
            self.update_dynamic_regex()
            
        elif choice == "Έξυπνη Μετατροπή":
            self.custom_regex_entry.configure(placeholder_text="π.χ. ΑΔΥΓΥΕΦΓ-33242432")
            self.custom_regex_var.set("ΑΔΥΓ-1234")
            self.update_dynamic_regex()

    def update_dynamic_regex(self, event=None):
        mode = self.regex_mode_var.get()
        user_input = self.custom_regex_var.get().strip()

        if mode == "Custom Regex (Για Προγραμματιστές)":
            self.final_regex = user_input
            self.active_regex_display.configure(text=f"Ενεργός Κανόνας: {self.final_regex}")
            
        elif mode == "Έξυπνη Μετατροπή":
            if not user_input:
                self.final_regex = ""
                self.active_regex_display.configure(text="Περιμένω παράδειγμα...")
                return

            parts = re.findall(r'[A-ZΑ-ΩΆΈΉΊΌΎΏ]+|[a-zα-ωάέήίόύώ]+|\d+|[^a-zA-ZΑ-Ωα-ω0-9\s]+', user_input)
            generated_regex = r"\b"
            
            for p in parts:
                if p.isupper():
                    generated_regex += f"[A-ZΑ-Ω]{{{len(p)}}}"
                elif p.islower():
                    generated_regex += f"[a-zα-ω]{{{len(p)}}}"
                elif p.isdigit():
                    generated_regex += f"\\d{{{len(p)}}}"
                else:
                    generated_regex += re.escape(p)
            
            generated_regex += r"\b"
            self.final_regex = generated_regex
            self.active_regex_display.configure(text=f"Αυτόματος Κανόνας: {self.final_regex}")

    # ΒΑΣΙΚΕΣ ΛΕΙΤΟΥΡΓΙΕΣ 
    def select_excel(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath: self.excel_path.set(filepath)

    def select_pdf(self):
        filepath = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if filepath: self.pdf_path.set(filepath)

    def load_data_to_cache(self):
        excel_file = self.excel_path.get()
        pdf_file = self.pdf_path.get()

        if not excel_file or not pdf_file:
            raise Exception("Παρακαλώ επιλέξτε και τα δύο αρχεία (Excel και PDF) πρώτα.")

        if not self.final_regex:
            raise Exception("Παρακαλώ εισάγετε έναν έγκυρο κανόνα ή παράδειγμα αναζήτησης.")

        self.root.after(0, lambda: self.status_label.configure(text="Διαβάζεται το Excel..."))
        
        self.cached_excel_codes = []
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                self.cached_excel_codes.append(str(row[0]).strip())
        
        pdf_text = ""
        with pdfplumber.open(pdf_file) as pdf:
            total_pages = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                self.root.after(0, lambda p=i+1, t=total_pages: self.status_label.configure(text=f"Διάβασμα PDF: Σελίδα {p} από {t}..."))
                text = page.extract_text()
                if text: 
                    pdf_text += text + "\n"
                time.sleep(0.05)

        self.root.after(0, lambda: self.status_label.configure(text="Αναζήτηση κωδικών στο κείμενο..."))
        
        try:
            self.cached_pdf_codes = re.findall(self.final_regex, pdf_text)
        except re.error:
            raise Exception("Προέκυψε συντακτικό λάθος στον κανόνα Regex. Παρακαλώ ελέγξτε τη μορφή.")

        self.last_loaded_excel = excel_file
        self.last_loaded_pdf = pdf_file

    def start_loading_ui(self):
        self.run_btn.configure(state="disabled")
        self.open_file_btn.pack_forget()
        self.progress.pack(pady=(5, 0))
        self.progress.start()
        self.status_label.pack(pady=(2, 0))
        
        self.preview_box.configure(state="normal")
        self.preview_box.delete("0.0", "end")
        self.preview_box.insert("end", "Γίνεται επεξεργασία, παρακαλώ περιμένετε...")
        self.preview_box.configure(state="disabled")

    def stop_loading_ui(self):
        self.progress.stop()
        self.progress.pack_forget()
        self.status_label.pack_forget()
        self.run_btn.configure(state="normal")

    def update_preview_ui(self, matches):
        self.preview_box.configure(state="normal")
        self.preview_box.delete("0.0", "end")
        
        self.preview_box.insert("end", f"{'Κωδικός Εργαζόμενου':<30} | {'Κατάσταση'}\n")
        self.preview_box.insert("end", "-" * 45 + "\n")
        
        if not matches:
            self.preview_box.insert("end", "Δεν βρέθηκαν κοινοί κωδικοί.\n")
        else:
            for match in matches:
                self.preview_box.insert("end", f"{match:<30} | Match\n")
                
        self.preview_box.configure(state="disabled")

    def start_matching_thread(self):
        self.start_loading_ui()
        threading.Thread(target=self.run_matching_logic, daemon=True).start()

    def run_matching_logic(self):
        try:
            self.load_data_to_cache()
            self.root.after(0, lambda: self.status_label.configure(text="Επεξεργασία δεδομένων..."))
            
            matches = [code for code in self.cached_excel_codes if code in self.cached_pdf_codes]

            self.root.after(0, lambda m=matches: self.update_preview_ui(m))

            if self.export_matches_var.get():
                wb_matches = openpyxl.Workbook()
                ws_matches = wb_matches.active
                ws_matches.title = "Matches"
                ws_matches.append(["Κωδικός Εργαζόμενου", "Κατάσταση"])
                for match in matches: ws_matches.append([match, "Match"])
                
                self.matches_filename = "Matches_Only.xlsx"
                wb_matches.save(self.matches_filename)
                self.root.after(0, lambda: self.open_file_btn.pack(pady=10))

            if self.highlight_excel_var.get():
                excel_file = self.excel_path.get()
                wb_original = openpyxl.load_workbook(excel_file)
                ws_original = wb_original.active
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                
                for row in ws_original.iter_rows(min_row=2):
                    cell = row[0]
                    if cell.value and str(cell.value).strip() in self.cached_pdf_codes:
                        cell.fill = green_fill
                
                wb_original.save(excel_file)

            self.root.after(0, lambda: messagebox.showinfo("Ολοκληρώθηκε", f"Η αντιστοίχιση ολοκληρώθηκε επιτυχώς!\nΒρέθηκαν {len(matches)} κοινοί κωδικοί."))

        except Exception as e:
            error_msg = str(e)
            if "Permission denied" in error_msg:
                error_msg = "Δεν ήταν δυνατή η αποθήκευση. Βεβαιωθείτε ότι το αρχείο Excel είναι ΚΛΕΙΣΤΟ και ξαναδοκιμάστε."
            self.root.after(0, lambda e=error_msg: messagebox.showerror("Σφάλμα", e))
            self.root.after(0, lambda: self.update_preview_ui([]))
        finally:
            self.root.after(0, self.stop_loading_ui)

    def open_matches_file(self):
        try:
            os.startfile(self.matches_filename)
        except Exception as e:
            messagebox.showerror("Σφάλμα", f"Δεν ήταν δυνατό το άνοιγμα:\n{str(e)}")

    def start_search_thread(self):
        code_to_search = self.search_code_var.get().strip()
        if not code_to_search:
            self.search_result_label.configure(text="Παρακαλώ γράψτε κωδικό.", text_color="red")
            return

        self.start_loading_ui()
        threading.Thread(target=self.run_search_logic, args=(code_to_search,), daemon=True).start()

    def run_search_logic(self, code_to_search):
        try:
            self.load_data_to_cache()
            
            in_excel = code_to_search in self.cached_excel_codes
            in_pdf = code_to_search in self.cached_pdf_codes

            if in_excel and in_pdf:
                self.root.after(0, lambda: self.search_result_label.configure(text="Ο εργαζόμενος υπάρχει και στα 2 αρχεία", text_color="green"))
            else:
                self.root.after(0, lambda: self.search_result_label.configure(text="Δεν υπάρχει και στα 2 αρχεία ο εργαζόμενος", text_color="red"))
                
        except Exception as e:
            self.root.after(0, lambda e=e: messagebox.showerror("Σφάλμα", str(e)))
        finally:
            self.root.after(0, self.stop_loading_ui)
            self.root.after(0, lambda: self.update_preview_ui([]))

if __name__ == "__main__":
    app_root = ctk.CTk()
    app = ModernDataMatcherApp(app_root)
    app_root.mainloop()